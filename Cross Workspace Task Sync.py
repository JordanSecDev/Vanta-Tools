#!/usr/bin/env python3
"""
Vanta cross-workspace security task deduplication

For each user found in multiple workspaces:
  - Identifies tasks they have COMPLETED in any workspace
  - Deactivates those same tasks in other workspaces where they are still
    pending/overdue (and not already disabled)
  - Sets deactivation reason: "completed in {workspace_name}"

Usage:
    # Dry run — no changes made, full report printed:
    python "Cross Workspace Task Sync.py" --config Workspaces.json --dry-run

    # Live run — executes deactivations:
    python "Cross Workspace Task Sync.py" --config Workspaces.json

    # Scope to specific task types or users:
    python "Cross Workspace Task Sync.py" --config Workspaces.json --dry-run \
        --task-type INSTALL_DEVICE_MONITORING \
        --email user@example.com

Refs:
  - Token:      https://api.vanta.com/oauth/token
  - People:     https://api.vanta.com/v1/people
  - Tests:      https://api.vanta.com/v1/tests
  - Entities:   https://api.vanta.com/v1/tests/{testId}/entities
  - Deactivate: https://api.vanta.com/v1/tests/{testId}/entities/{entityId}/deactivate
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple

import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

TOKEN_URL       = "https://api.vanta.com/oauth/token"
PEOPLE_URL      = "https://api.vanta.com/v1/people"
TESTS_URL       = "https://api.vanta.com/v1/tests"
ENTITIES_URL    = "https://api.vanta.com/v1/tests/{testId}/entities"
DEACTIVATE_URL  = "https://api.vanta.com/v1/tests/{testId}/entities/{entityId}/deactivate"

READ_SCOPE  = "vanta-api.all:read"
WRITE_SCOPE = "vanta-api.all:write"

# Maps the /v1/people tasksSummary.details key for each task-type enum value.
TASK_DETAIL_KEY: Dict[str, str] = {
    "INSTALL_DEVICE_MONITORING":        "installDeviceMonitoring",
    "COMPLETE_TRAININGS":               "completeTrainings",
    "ACCEPT_POLICIES":                  "acceptPolicies",
    "COMPLETE_BACKGROUND_CHECKS":       "completeBackgroundChecks",
    "COMPLETE_CUSTOM_TASKS":            "completeCustomTasks",
    "COMPLETE_CUSTOM_OFFBOARDING_TASKS":"completeCustomOffboardingTasks",
}

# Keywords used to match GET /v1/tests response names to task types.
# ACCEPT_POLICIES and COMPLETE_TRAININGS map to MULTIPLE tests (one per policy/training
# programme) — ALL matching tests are collected and each is deactivated independently.
TASK_KEYWORDS: Dict[str, List[str]] = {
    "INSTALL_DEVICE_MONITORING":        ["device monitoring", "install device", "mdm", "agent install",
                                         "computers monitored", "employees-without-laptops",
                                         "without laptops"],
    "COMPLETE_TRAININGS":               ["training", "security awareness"],
    "ACCEPT_POLICIES":                  ["personnel agree", "accept polic", "policy acceptance",
                                         "acceptable use", "policy sign"],
    "COMPLETE_BACKGROUND_CHECKS":       ["background check", "background screening"],
    "COMPLETE_CUSTOM_TASKS":            ["custom task"],
    "COMPLETE_CUSTOM_OFFBOARDING_TASKS":["offboard"],
}

# Task types that fan out to many tests (one per policy/training).
# For these, ALL matching tests are collected rather than just the first.
MULTI_TEST_TASK_TYPES: Set[str] = {"ACCEPT_POLICIES", "COMPLETE_TRAININGS"}

# Task statuses that we treat as "needs deactivation" if completed elsewhere.
PENDING_STATUSES: Set[str] = {"OVERDUE", "DUE_SOON", "NONE"}


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class Workspace:
    name: str
    client_id: str
    client_secret: str


@dataclass
class PersonTask:
    workspace:       str
    person_id:       str
    email:           str
    display_name:    str
    task_type:       str
    status:          str            # COMPLETE | OVERDUE | DUE_SOON | NONE
    disabled:        bool
    disabled_reason: Optional[str]  # reason stored in Vanta when disabled


@dataclass
class CompletedDeactivation:
    email:                  str
    display_name:           str
    task_type:              str
    deactivated_workspace:  str           # workspace where task is disabled
    completed_in_workspace: str           # workspace where task is COMPLETE
    disabled_reason:        Optional[str] # reason currently stored in Vanta


@dataclass
class DeactivationCandidate:
    email:                   str
    display_name:            str
    task_type:               str
    target_workspace:        str
    target_person_id:        str
    completed_in_workspace:  str
    # test_ids: all tests to deactivate for this candidate (one per policy/training,
    # or a single entry for INSTALL_DEVICE_MONITORING / COMPLETE_BACKGROUND_CHECKS).
    # entity_id is resolved at runtime; falls back to target_person_id if not found.
    test_ids:  List[str] = None   # type: ignore[assignment]
    entity_id: Optional[str] = None

    def __post_init__(self):
        if self.test_ids is None:
            self.test_ids = []


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

def load_workspaces(path: str) -> List[Workspace]:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, dict) or "workspaces" not in data:
        raise ValueError("Config must be a JSON object with key 'workspaces' (array).")
    return [
        Workspace(name=w["name"], client_id=w["client_id"], client_secret=w["client_secret"])
        for w in data["workspaces"]
    ]


def get_token(ws: Workspace, scope: str) -> str:
    r = requests.post(TOKEN_URL, json={
        "client_id":     ws.client_id,
        "client_secret": ws.client_secret,
        "scope":         scope,
        "grant_type":    "client_credentials",
    }, timeout=30)
    if r.status_code != 200:
        raise RuntimeError(f"[{ws.name}] Token request failed: {r.status_code} {r.text}")
    return r.json()["access_token"]


# ---------------------------------------------------------------------------
# HTTP helpers
# ---------------------------------------------------------------------------

def http_get(url: str, token: str, params: Dict[str, Any] = None) -> Dict[str, Any]:
    headers = {"Accept": "application/json", "Authorization": f"Bearer {token}"}
    r = requests.get(url, headers=headers, params=params or {}, timeout=60)
    if r.status_code != 200:
        raise RuntimeError(f"GET {url} failed: {r.status_code} {r.text}")
    return r.json()


def http_post(url: str, token: str, body: Dict[str, Any] = None,
              max_retries: int = 5) -> Dict[str, Any]:
    headers = {
        "Accept":        "application/json",
        "Content-Type":  "application/json",
        "Authorization": f"Bearer {token}",
    }
    backoff = 2.0
    for attempt in range(max_retries):
        r = requests.post(url, headers=headers, json=body or {}, timeout=60)
        if r.status_code == 429:
            wait = float(r.headers.get("Retry-After", backoff))
            time.sleep(wait)
            backoff = min(backoff * 2, 60)
            continue
        if r.status_code not in (200, 201, 202):
            raise RuntimeError(f"POST {url} failed: {r.status_code} {r.text}")
        return r.json()
    raise RuntimeError(f"POST {url} failed: still 429 after {max_retries} retries")


# ---------------------------------------------------------------------------
# Generic paginator — handles both response shapes seen in the Vanta API
# ---------------------------------------------------------------------------

def iter_pages(url: str, token: str, params: Dict[str, Any], page_size: int = 100) -> Iterable[Dict[str, Any]]:
    """
    Iterates a paginated Vanta list endpoint.
    Handles response shapes:
      A) {"results": [...], "pageInfo": {...}}
      B) {"results": {"data": [...], "pageInfo": {...}}}
      C) {"results": {"nodes": [...], "pageInfo": {...}}}
    """
    cursor: Optional[str] = None

    while True:
        p = dict(params)
        p["pageSize"] = page_size
        if cursor:
            p["pageCursor"] = cursor

        data = http_get(url, token, p)
        results = data.get("results", {})

        if isinstance(results, list):
            nodes     = results
            page_info = data.get("pageInfo") or data.get("resultsPageInfo") or {}
        elif isinstance(results, dict):
            nodes     = results.get("data") or results.get("nodes") or results.get("results") or []
            page_info = results.get("pageInfo") or {}
        else:
            nodes     = []
            page_info = {}

        if not nodes:
            break

        for node in nodes:
            yield node

        has_next   = bool(page_info.get("hasNextPage"))
        end_cursor = page_info.get("endCursor")

        # Some endpoints use nextPageCursor at the top level
        if not has_next and data.get("nextPageCursor"):
            has_next   = True
            end_cursor = data["nextPageCursor"]

        if not has_next or not end_cursor:
            break
        cursor = end_cursor


# ---------------------------------------------------------------------------
# Phase 1 — Collect people + task statuses
# ---------------------------------------------------------------------------

def safe_get(d: Any, path: List[str], default=None) -> Any:
    cur = d
    for k in path:
        if not isinstance(cur, dict) or k not in cur:
            return default
        cur = cur[k]
    return cur


def parse_disabled(val: Any) -> Tuple[bool, Optional[str]]:
    """Returns (is_disabled, reason_or_None)."""
    if isinstance(val, bool):
        return val, None
    if isinstance(val, dict):
        return True, val.get("reason") or None
    return False, None


def fetch_people_tasks(ws: Workspace, token: str, debug_fn) -> List[PersonTask]:
    rows: List[PersonTask] = []

    for person in iter_pages(PEOPLE_URL, token, {}):
        email        = (person.get("emailAddress") or "").strip().lower()
        person_id    = person.get("id", "")
        display_name = safe_get(person, ["name", "display"]) or email
        details      = safe_get(person, ["tasksSummary", "details"]) or {}

        for task_type, detail_key in TASK_DETAIL_KEY.items():
            task = details.get(detail_key) or {}
            if not task:
                continue
            status          = task.get("status") or "NONE"
            disabled, reason = parse_disabled(task.get("disabled"))

            debug_fn(f"  [{ws.name}] {email} | {task_type} | status={status} disabled={disabled}")

            rows.append(PersonTask(
                workspace       = ws.name,
                person_id       = person_id,
                email           = email,
                display_name    = display_name,
                task_type       = task_type,
                status          = status,
                disabled        = disabled,
                disabled_reason = reason,
            ))

    return rows


# ---------------------------------------------------------------------------
# Phase 2 — Plan deactivations
# ---------------------------------------------------------------------------

def plan_deactivations(all_tasks: List[PersonTask]) -> List[DeactivationCandidate]:
    """
    For each (email, task_type): if COMPLETE in ≥1 workspace and pending in
    another (and not already disabled), create a DeactivationCandidate.
    """
    # email → task_type → [PersonTask]
    index: Dict[str, Dict[str, List[PersonTask]]] = {}
    for t in all_tasks:
        index.setdefault(t.email, {}).setdefault(t.task_type, []).append(t)

    candidates: List[DeactivationCandidate] = []

    for email, task_map in index.items():
        for task_type, tasks in task_map.items():
            completed = [t for t in tasks if t.status == "COMPLETE"]
            pending   = [t for t in tasks if t.status in PENDING_STATUSES and not t.disabled]

            if not completed or not pending:
                continue

            # Prefer the workspace with the earliest completion (stable sort)
            source = completed[0]

            for p in pending:
                if p.workspace == source.workspace:
                    continue  # shouldn't happen but guard
                candidates.append(DeactivationCandidate(
                    email                  = email,
                    display_name           = p.display_name,
                    task_type              = task_type,
                    target_workspace       = p.workspace,
                    target_person_id       = p.person_id,
                    completed_in_workspace = source.workspace,
                ))

    return candidates


def find_completed_deactivations(all_tasks: List[PersonTask]) -> List[CompletedDeactivation]:
    """
    For each (email, task_type): if COMPLETE in ≥1 workspace and disabled in
    another workspace, that disabled entry is a previously completed deactivation.
    """
    index: Dict[str, Dict[str, List[PersonTask]]] = {}
    for t in all_tasks:
        index.setdefault(t.email, {}).setdefault(t.task_type, []).append(t)

    completed_deactivations: List[CompletedDeactivation] = []

    for email, task_map in index.items():
        for task_type, tasks in task_map.items():
            completed = [t for t in tasks if t.status == "COMPLETE"]
            disabled  = [t for t in tasks if t.disabled]

            if not completed or not disabled:
                continue

            source = completed[0]
            for d in disabled:
                if d.workspace == source.workspace:
                    continue
                completed_deactivations.append(CompletedDeactivation(
                    email                  = email,
                    display_name           = d.display_name,
                    task_type              = task_type,
                    deactivated_workspace  = d.workspace,
                    completed_in_workspace = source.workspace,
                    disabled_reason        = d.disabled_reason,
                ))

    return completed_deactivations


# ---------------------------------------------------------------------------
# Phase 3 — Resolve test IDs per workspace
# ---------------------------------------------------------------------------

def match_task_type(test_name: str) -> Optional[str]:
    name_lower = test_name.lower()
    for task_type, keywords in TASK_KEYWORDS.items():
        if any(kw in name_lower for kw in keywords):
            return task_type
    return None


def fetch_test_id_map(ws: Workspace, token: str, debug_fn,
                      framework: Optional[str] = None) -> Dict[str, List[str]]:
    """
    Returns {task_type: [testId, ...]} for all tests in this workspace.

    For MULTI_TEST_TASK_TYPES (ACCEPT_POLICIES, COMPLETE_TRAININGS) every
    matching test is collected because each policy / training programme has
    its own test.  For other task types only the first match is kept.

    If framework is set (e.g. "soc2"), only tests belonging to that framework
    are fetched via the frameworkFilter query parameter.
    """
    mapping: Dict[str, List[str]] = {}
    params: Dict[str, Any] = {}
    if framework:
        params["frameworkFilter"] = framework

    for test in iter_pages(TESTS_URL, token, params):
        name    = test.get("name") or ""
        test_id = test.get("id")
        if not test_id:
            continue
        task_type = match_task_type(name)
        debug_fn(f"  [{ws.name}] test '{name}' → matched={task_type}")
        if not task_type:
            continue
        if task_type in MULTI_TEST_TASK_TYPES:
            mapping.setdefault(task_type, []).append(test_id)
        elif task_type not in mapping:
            mapping[task_type] = [test_id]   # single-entry list for consistency

    return mapping


# ---------------------------------------------------------------------------
# Phase 4 — Assign entity IDs
# ---------------------------------------------------------------------------
# For all person-level security tasks (trainings, policies, device monitoring,
# background checks) Vanta uses the person's own ID as the test entity ID.
# We therefore set entity_id = target_person_id directly rather than scanning
# the (potentially large) entities endpoint.  The live run will confirm this:
# a 200 SUCCESS means it matched; an error will surface the mismatch.


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

def print_dry_run_report(candidates: List[DeactivationCandidate],
                         completed: List[CompletedDeactivation]) -> None:
    W = 80
    print()
    print("=" * W)
    print("DRY RUN REPORT — No changes have been made")
    print("=" * W)

    if not candidates:
        print("\nNo deactivation candidates found.")
        print("  This means no user has the same task COMPLETE in one workspace")
        print("  and still pending in another.")
        return

    by_workspace: Dict[str, List[DeactivationCandidate]] = {}
    for c in candidates:
        by_workspace.setdefault(c.target_workspace, []).append(c)

    total_api_calls = unresolved_test = 0

    for ws_name in sorted(by_workspace):
        ws_candidates = by_workspace[ws_name]
        print(f"\nWorkspace: {ws_name}  ({len(ws_candidates)} candidate(s))")
        print("-" * W)

        for c in sorted(ws_candidates, key=lambda x: (x.task_type, x.email)):
            test_ids  = c.test_ids
            entity_id = c.entity_id   # always set to person_id by Phase 4

            if test_ids:
                tag = "[READY]          "
                total_api_calls += len(test_ids)
            else:
                tag = "[TEST_UNRESOLVED]"
                unresolved_test += 1

            reason = f"completed in {c.completed_in_workspace}"

            print(f"  {tag}")
            print(f"    User      : {c.display_name} <{c.email}>")
            print(f"    Task      : {c.task_type}")
            print(f"    Deactivate: '{c.target_workspace}'")
            print(f"    Reason    : \"{reason}\"")
            print(f"    Entity ID : {entity_id}")
            if test_ids:
                for tid in test_ids:
                    url = DEACTIVATE_URL.format(testId=tid, entityId=entity_id)
                    print(f"    API call  : POST {url}")
            else:
                print(f"    API call  : N/A — test ID not resolved (check TASK_KEYWORDS)")
            print()

    total = len(candidates)
    print("=" * W)
    print(f"SUMMARY")
    print(f"  Unique candidates : {total}")
    print(f"  Total API calls   : {total_api_calls}  (candidates × tests per task type)")
    print(f"  Test unresolved   : {unresolved_test}  (no test name matched TASK_KEYWORDS)")
    print("=" * W)

    if unresolved_test:
        print()
        print("NOTE: Unresolved test IDs mean no test name matched the TASK_KEYWORDS patterns.")
        print("      Run with --debug to see all test names returned by GET /v1/tests.")
        print("      Add the matching keyword to TASK_KEYWORDS in the script and re-run.")

    # ------------------------------------------------------------------
    # Section 2: Completed deactivations
    # ------------------------------------------------------------------
    W = 80
    print()
    print("=" * W)
    print("COMPLETED DEACTIVATIONS — already disabled in Vanta")
    print("=" * W)

    if not completed:
        print("\nNone found.")
    else:
        by_workspace: Dict[str, List[CompletedDeactivation]] = {}
        for c in completed:
            by_workspace.setdefault(c.deactivated_workspace, []).append(c)

        for ws_name in sorted(by_workspace):
            ws_items = by_workspace[ws_name]
            print(f"\nWorkspace: {ws_name}  ({len(ws_items)} item(s))")
            print("-" * W)
            for c in sorted(ws_items, key=lambda x: (x.task_type, x.email)):
                print(f"  [DONE]")
                print(f"    User              : {c.display_name} <{c.email}>")
                print(f"    Task              : {c.task_type}")
                print(f"    Deactivated in    : '{c.deactivated_workspace}'")
                print(f"    Completed in      : '{c.completed_in_workspace}'")
                print(f"    Stored reason     : {c.disabled_reason or '(none)'}")
                print()

        print("=" * W)
        print(f"TOTAL COMPLETED: {len(completed)}")
        print("=" * W)


# ---------------------------------------------------------------------------
# Phase 5 — Execute deactivations
# ---------------------------------------------------------------------------

def do_deactivate(
    candidate: DeactivationCandidate,
    test_id:   str,
    entity_id: str,
    token:     str,
) -> Tuple[bool, str, bool]:
    """Returns (ok, message, skipped).  skipped=True means 404 — already resolved."""
    url    = DEACTIVATE_URL.format(testId=test_id, entityId=entity_id)
    reason = f"completed in {candidate.completed_in_workspace}"
    try:
        result = http_post(url, token, {"deactivateReason": reason})
        status = result.get("status", "UNKNOWN")
        return status == "SUCCESS", f"status={status} id={result.get('id')} reason_sent='{reason}'", False
    except RuntimeError as e:
        if " 404 " in str(e) or str(e).startswith("POST") and "404" in str(e):
            return True, f"SKIPPED — entity not found (404), already resolved or never failing", True
        return False, str(e), False
    except Exception as e:
        return False, str(e), False


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

# Header style constants
_HDR_FONT    = Font(bold=True, color="FFFFFF")
_HDR_FILL    = PatternFill("solid", fgColor="1F4E79")
_READY_FILL  = PatternFill("solid", fgColor="C6EFCE")   # green
_SKIP_FILL   = PatternFill("solid", fgColor="FFCCCC")   # red
_WRAP        = Alignment(wrap_text=True, vertical="top")


def _set_header_row(ws_sheet, headers: List[str]) -> None:
    ws_sheet.append(headers)
    for cell in ws_sheet[1]:
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = _WRAP
    ws_sheet.freeze_panes = "A2"


def _autofit(ws_sheet, max_width: int = 60) -> None:
    for col_cells in ws_sheet.columns:
        length = max(len(str(c.value or "")) for c in col_cells)
        ws_sheet.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 2, max_width)


def write_plan_excel(candidates: List[DeactivationCandidate],
                     completed: List[CompletedDeactivation],
                     path: str) -> int:
    """
    Writes the deactivation plan to an Excel workbook with two sheets:
      - Plan       : one row per API call (candidate × test_id)
      - Summary    : candidate and API-call counts grouped by workspace and task type
    Returns the total number of rows written to the Plan sheet.
    """
    wb = Workbook()

    # ------------------------------------------------------------------
    # Sheet 1: Plan — one row per deactivation API call
    # ------------------------------------------------------------------
    ws_plan = wb.active
    ws_plan.title = "Deactivation Plan"

    plan_headers = [
        "Status",
        "User Email",
        "Display Name",
        "Task Type",
        "Target Workspace",
        "Completed In Workspace",
        "Reason",
        "Test ID",
        "Entity ID",
        "API URL",
    ]
    _set_header_row(ws_plan, plan_headers)

    plan_rows = 0
    for c in sorted(candidates, key=lambda x: (x.target_workspace, x.task_type, x.email)):
        reason = f"completed in {c.completed_in_workspace}"
        if c.test_ids:
            for tid in c.test_ids:
                url = DEACTIVATE_URL.format(testId=tid, entityId=c.entity_id)
                row = ["READY", c.email, c.display_name, c.task_type,
                       c.target_workspace, c.completed_in_workspace,
                       reason, tid, c.entity_id, url]
                ws_plan.append(row)
                for cell in ws_plan[ws_plan.max_row]:
                    cell.alignment = _WRAP
                ws_plan[ws_plan.max_row][0].fill = _READY_FILL
                plan_rows += 1
        else:
            row = ["TEST_UNRESOLVED", c.email, c.display_name, c.task_type,
                   c.target_workspace, c.completed_in_workspace,
                   reason, "NOT FOUND", c.entity_id, "N/A"]
            ws_plan.append(row)
            for cell in ws_plan[ws_plan.max_row]:
                cell.alignment = _WRAP
            ws_plan[ws_plan.max_row][0].fill = _SKIP_FILL

    _autofit(ws_plan)

    # ------------------------------------------------------------------
    # Sheet 2: Summary — counts by workspace × task type
    # ------------------------------------------------------------------
    ws_sum = wb.create_sheet("Summary")
    sum_headers = ["Target Workspace", "Task Type", "Candidates", "API Calls (tests × users)", "Status"]
    _set_header_row(ws_sum, sum_headers)

    # Build counts
    from collections import defaultdict
    counts: Dict[tuple, Dict[str, int]] = defaultdict(lambda: {"candidates": 0, "api_calls": 0, "unresolved": 0})
    for c in candidates:
        key = (c.target_workspace, c.task_type)
        counts[key]["candidates"] += 1
        if c.test_ids:
            counts[key]["api_calls"] += len(c.test_ids)
        else:
            counts[key]["unresolved"] += 1

    for (ws_name, task_type) in sorted(counts.keys()):
        d = counts[(ws_name, task_type)]
        unresolved = d["unresolved"]
        status = "TEST_UNRESOLVED" if unresolved == d["candidates"] else (
                 "PARTIAL" if unresolved > 0 else "READY")
        row = [ws_name, task_type, d["candidates"], d["api_calls"], status]
        ws_sum.append(row)
        fill = _READY_FILL if status == "READY" else _SKIP_FILL
        ws_sum[ws_sum.max_row][4].fill = fill

    # Totals row
    total_candidates = len(candidates)
    total_api_calls  = sum(d["api_calls"] for d in counts.values())
    ws_sum.append([])
    ws_sum.append(["TOTAL", "", total_candidates, total_api_calls, ""])
    for cell in ws_sum[ws_sum.max_row]:
        cell.font = Font(bold=True)

    _autofit(ws_sum)

    # ------------------------------------------------------------------
    # Sheet 3: Completed Deactivations
    # ------------------------------------------------------------------
    ws_done = wb.create_sheet("Completed Deactivations")
    done_headers = ["User Email", "Display Name", "Task Type",
                    "Deactivated Workspace", "Completed In Workspace", "Stored Reason"]
    _set_header_row(ws_done, done_headers)

    for c in sorted(completed, key=lambda x: (x.deactivated_workspace, x.task_type, x.email)):
        ws_done.append([
            c.email, c.display_name, c.task_type,
            c.deactivated_workspace, c.completed_in_workspace,
            c.disabled_reason or "",
        ])
        for cell in ws_done[ws_done.max_row]:
            cell.alignment = _WRAP
        ws_done[ws_done.max_row][0].fill = _READY_FILL

    _autofit(ws_done)

    wb.save(path)
    return plan_rows


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(description="Vanta cross-workspace task deduplication")
    ap.add_argument("--config",     required=True,        help="Path to workspaces JSON config")
    ap.add_argument("--dry-run",    action="store_true",  help="Report only — no changes made")
    ap.add_argument("--out-prefix", default="vanta_task_sync_plan",
                                                          help="Output filename prefix for dry-run Excel (default: vanta_task_sync_plan)")
    ap.add_argument("--task-type",  action="append", default=[], help="Limit to task type(s) (repeatable)")
    ap.add_argument("--email",      action="append", default=[], help="Limit to user email(s) (repeatable)")
    ap.add_argument("--workers",    type=int, default=5,         help="Max parallel deactivation threads (default: 5)")
    ap.add_argument("--framework",  default=None,                help="Restrict to a Vanta framework slug, e.g. soc2")
    ap.add_argument("--debug",      action="store_true",  help="Verbose debug output to stderr")
    args = ap.parse_args()

    def log(msg: str)   -> None: print(msg, file=sys.stderr)
    def debug(msg: str) -> None:
        if args.debug:
            print(f"[DEBUG] {msg}", file=sys.stderr)

    workspaces   = load_workspaces(args.config)
    task_filter  = {t.upper() for t in args.task_type}
    email_filter = {e.strip().lower() for e in args.email}

    # -----------------------------------------------------------------------
    # Phase 1: Collect people + task statuses (parallel per workspace)
    # -----------------------------------------------------------------------
    log("\n=== Phase 1: Collecting people and task statuses ===")
    all_tasks:   List[PersonTask] = []
    read_tokens: Dict[str, str]   = {}
    _lock = threading.Lock()

    def _fetch_workspace(ws: Workspace):
        log(f"[{ws.name}] Authenticating (read)…")
        try:
            token = get_token(ws, READ_SCOPE)
        except Exception as e:
            log(f"[{ws.name}] Auth FAILED: {e}")
            return
        log(f"[{ws.name}] Fetching people…")
        try:
            tasks = fetch_people_tasks(ws, token, debug)
            if email_filter:
                tasks = [t for t in tasks if t.email in email_filter]
            if task_filter:
                tasks = [t for t in tasks if t.task_type in task_filter]
            log(f"[{ws.name}] {len(tasks)} task record(s) collected.")
            with _lock:
                read_tokens[ws.name] = token
                all_tasks.extend(tasks)
        except Exception as e:
            log(f"[{ws.name}] People fetch FAILED: {e}")

    with ThreadPoolExecutor(max_workers=len(workspaces)) as pool:
        list(pool.map(_fetch_workspace, workspaces))

    if not all_tasks:
        log("No task data collected. Check your config and credentials.")
        return 1

    # -----------------------------------------------------------------------
    # Phase 2: Plan deactivations
    # -----------------------------------------------------------------------
    log("\n=== Phase 2: Planning deactivations ===")
    candidates = plan_deactivations(all_tasks)
    completed  = find_completed_deactivations(all_tasks)
    log(f"Found {len(candidates)} outstanding candidate(s), {len(completed)} already completed.")

    if not candidates:
        log("Nothing to deactivate — no users share an incomplete task across workspaces.")
        return 0

    # -----------------------------------------------------------------------
    # Phase 3: Resolve test IDs per target workspace (parallel)
    # -----------------------------------------------------------------------
    log("\n=== Phase 3: Resolving test IDs ===")
    target_workspaces = {c.target_workspace for c in candidates}
    test_maps: Dict[str, Dict[str, List[str]]] = {}

    def _fetch_tests(ws: Workspace):
        token = read_tokens.get(ws.name)
        if not token:
            return
        log(f"[{ws.name}] Fetching tests…")
        try:
            result = fetch_test_id_map(ws, token, debug, framework=args.framework)
            total_ids = sum(len(v) for v in result.values())
            log(f"[{ws.name}] Matched {len(result)} task type(s), {total_ids} test(s) total.")
            with _lock:
                test_maps[ws.name] = result
        except Exception as e:
            log(f"[{ws.name}] Tests fetch FAILED: {e}")

    target_ws_objs = [ws for ws in workspaces if ws.name in target_workspaces]
    with ThreadPoolExecutor(max_workers=len(target_ws_objs) or 1) as pool:
        list(pool.map(_fetch_tests, target_ws_objs))

    # Attach resolved test ID lists to each candidate.
    for c in candidates:
        c.test_ids = test_maps.get(c.target_workspace, {}).get(c.task_type, [])

    # -----------------------------------------------------------------------
    # Phase 4: Assign entity IDs (person ID = entity ID for all person-level tasks)
    # -----------------------------------------------------------------------
    log("\n=== Phase 4: Assigning entity IDs ===")
    for c in candidates:
        c.entity_id = f"User-{c.target_person_id}"
        debug(f"  entity_id={c.entity_id} for {c.email} in {c.target_workspace}")

    # -----------------------------------------------------------------------
    # Phase 5: Report (dry run) or Execute
    # -----------------------------------------------------------------------
    if args.dry_run:
        print_dry_run_report(candidates, completed)
        xlsx_path = f"{args.out_prefix}.xlsx"
        n = write_plan_excel(candidates, completed, xlsx_path)
        log(f"\nExcel plan written → {xlsx_path}  ({n} outstanding rows, {len(completed)} completed rows)")
        return 0

    log("\n=== Phase 5: Executing deactivations ===")

    # Acquire write tokens for target workspaces
    write_tokens: Dict[str, str] = {}
    for ws in workspaces:
        if ws.name not in target_workspaces:
            continue
        log(f"[{ws.name}] Authenticating (write)…")
        try:
            write_tokens[ws.name] = get_token(ws, WRITE_SCOPE)
        except Exception as e:
            log(f"[{ws.name}] Write auth FAILED: {e} — skipping candidates for this workspace.")

    success = failure = skipped = already_resolved = 0
    _counters_lock = threading.Lock()

    # Build the flat list of (candidate, test_id) work items upfront,
    # filtering out anything without a token or test IDs.
    work_items = []
    for c in candidates:
        token = write_tokens.get(c.target_workspace)
        if not c.test_ids:
            log(f"SKIP [{c.target_workspace}] {c.email} / {c.task_type} — no test IDs resolved")
            with _counters_lock:
                skipped += 1
            continue
        if not token:
            log(f"SKIP [{c.target_workspace}] {c.email} / {c.task_type} — no write token")
            with _counters_lock:
                skipped += 1
            continue
        for test_id in c.test_ids:
            work_items.append((c, test_id, token))

    log(f"Dispatching {len(work_items)} deactivation call(s) across {args.workers} worker(s)…")

    def _deactivate_one(item):
        c, test_id, token = item
        ok, msg, was_404 = do_deactivate(c, test_id, c.entity_id, token)
        return c, test_id, ok, msg, was_404

    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        for c, test_id, ok, msg, was_404 in pool.map(_deactivate_one, work_items):
            if was_404:
                tag = "404"
                with _counters_lock:
                    already_resolved += 1
            elif ok:
                tag = "OK "
                with _counters_lock:
                    success += 1
            else:
                tag = "ERR"
                with _counters_lock:
                    failure += 1
            log(f"[{tag}] [{c.target_workspace}] {c.email} / {c.task_type} / {test_id} — {msg}")

    log(f"\nDone.  success={success}  failure={failure}  skipped={skipped}  already_resolved(404)={already_resolved}")
    return 0 if failure == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
